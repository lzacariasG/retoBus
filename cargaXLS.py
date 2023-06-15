import psycopg2 
import openpyxl
from datetime import date
from fastapi import FastAPI

api = FastAPI()

@api.get("/cargaExcel")


def validaNombre( nombreA: str):
    if nombreA.endswith(".xlsx"):
        return True
    else:
        return False


def lee_Excel():
    self.archivoExcel = openpyxl.load_workbook(self.nombreA)
    hoja = self.archivoExcel.active
    if hoja.max_column != 6:
        return False
    else:
        celdaFin = 'F'+ str(hoja.max_row -1)
        multiple_cells = hoja['A2': celdaFin]
        tabla = [] 
        for row in multiple_cells:
            lineaV = []
            for cell in row:
                lineaV.append(str(cell.value))
            tabla.append(lineaV)
        self.tuplas = tuple(tabla)



def registraArchivo(self):
    fechaC = date.today()
    sqlC = "INSERT INTO archivoscargados(id, nombre, fchCarga) SELECT nextval('archivo_id_seq'),'"+self.nombreA+"', '"+str(fechaC)+"'"
    self.conection = psycopg2.connect("dbname=bdBous user=jonathan password=admin")
    self.cursor = self.conection.cursor()
    self.cursor.execute (sqlC)
    self.conection.commit()
    self.cursor.close()
    self.conection.close()


def insertBD(self):
    self.conection = psycopg2.connect("dbname=bdBous user=jonathan password=admin")
    self.cursor = self.conection.cursor()
    self.cursor.executemany ('INSERT INTO archivoTmp(cliente, contrato, fch_compra, ciudad, empresa, valoradeudo, idArchivo) SELECT %s,%s,%s,%s, %s, %s, a.id FROM archivoscargados a where a.nombre = \''+self.nombreA+'\'', self.tuplas)
    self.conection.commit()
    self.conection.close()


def actualizaEst(self):
    fechaC = date.today()
    sqlC = "SELECT cargaRegistros('"+self.nombreA+"', '"+str(fechaC)+"')"
    self.conection = psycopg2.connect("dbname=bdBous user=jonathan password=admin")
    self.cursor = self.conection.cursor()
    self.cursor.execute (sqlC)
    self.conection.commit()
    self.cursor.close()
    self.conection.close()

    
        

instC = import_in_db()
instC.nombreA = "cargaBD.xlsx"
if instC.validaNombre():
    instC.lee_Excel()
    instC.registraArchivo()
    instC.insertBD()
    instC.actualizaEst()
    